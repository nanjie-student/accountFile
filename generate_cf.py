from openpyxl import Workbook 
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side 
from openpyxl.utils import get_column_letter 

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "现金流量表"

# 设置表头
ws['A1'] = "[公司全称]"
ws['A2'] = "现金流量表"
ws['A3'] = "截至2025年12月31日止年度"

# 合并表头
ws.merge_cells('A1:C1')
ws.merge_cells('A2:C2')
ws.merge_cells('A3:C3')

# 设置表头样式
header_font = Font(name='微软雅黑', size=14, bold=True)
for cell in ws['A1:C3']:
    for c in cell:
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center')

# 设置列宽
ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15

# 定义项目列表
items = [
    ("一、经营活动产生的现金流量", None, None),
    ("销售商品、提供劳务收到的现金", "C5", "D5"),
    ("收到的税费返还", "C6", "D6"),
    ("收到其他与经营活动有关的现金", "C7", "D7"),
    ("经营活动现金流入小计", "C8", "D8"),
    ("购买商品、接受劳务支付的现金", "C11", "D11"),
    ("支付给职工以及为职工支付的现金", "C12", "D12"),
    ("支付的各项税费", "C13", "D13"),
    ("支付其他与经营活动有关的现金", "C14", "D14"),
    ("经营活动现金流出小计", "C15", "D15"),
    ("经营活动产生的现金流量净额", "C16", "D16"),
    ("二、投资活动产生的现金流量", None, None),
    ("收回投资收到的现金", "C19", "D19"),
    ("取得投资收益收到的现金", "C20", "D20"),
    ("处置固定资产、无形资产和其他长期资产收回的现金净额", "C21", "D21"),
    ("处置子公司及其他营业单位收到的现金净额", "C22", "D22"),
    ("收到其他与投资活动有关的现金", "C23", "D23"),
    ("投资活动现金流入小计", "C24", "D24"),
    ("购建固定资产、无形资产和其他长期资产支付的现金", "C27", "D27"),
    ("投资支付的现金", "C28", "D28"),
    ("取得子公司及其他营业单位支付的现金净额", "C29", "D29"),
    ("支付其他与投资活动有关的现金", "C30", "D30"),
    ("投资活动现金流出小计", "C31", "D31"),
    ("投资活动产生的现金流量净额", "C32", "D32"),
    ("三、筹资活动产生的现金流量", None, None),
    ("吸收投资收到的现金", "C35", "D35"),
    ("取得借款收到的现金", "C36", "D36"),
    ("发行债券收到的现金", "C37", "D37"),
    ("收到其他与筹资活动有关的现金", "C38", "D38"),
    ("筹资活动现金流入小计", "C39", "D39"),
    ("偿还债务支付的现金", "C42", "D42"),
    ("分配股利、利润或偿付利息支付的现金", "C43", "D43"),
    ("支付其他与筹资活动有关的现金", "C44", "D44"),
    ("筹资活动现金流出小计", "C45", "D45"),
    ("筹资活动产生的现金流量净额", "C46", "D46"),
    ("四、汇率变动对现金及现金等价物的影响", "C47", "D47"),
    ("五、现金及现金等价物净增加额", "C48", "D48"),
    ("加：期初现金及现金等价物余额", "C49", "D49"),
    ("六、期末现金及现金等价物余额", "C50", "D50")
]

# 填充项目
row = 5
for item, col_b, col_c in items:
    if item is None:
        row += 1
        continue
    ws[f'A{row}'] = item
    if col_b:
        ws[col_b] = ""
        ws[col_c] = ""
        # 设置公式
        if "小计" in item:
            start_row = row - 3 if "流入" in item else row - 4
            end_row = row - 1
            formula = f"=SUM({get_column_letter(ord(col_b) - ord('A') + 1)}{start_row}:{get_column_letter(ord(col_b) - ord('A') + 1)}{end_row})"
            ws[col_b] = formula
            ws[col_c] = formula.replace(col_b, col_c)
        elif "净额" in item:
            if "经营" in item:
                inflow_cell = f"{get_column_letter(ord(col_b) - ord('A') + 1)}{row - 8}"
                outflow_cell = f"{get_column_letter(ord(col_b) - ord('A') + 1)}{row - 1}"
                ws[col_b] = f"={inflow_cell}-{outflow_cell}"
                ws[col_c] = f"={inflow_cell.replace(col_b, col_c)}-{outflow_cell.replace(col_b, col_c)}"
            elif "投资" in item:
                inflow_cell = f"{get_column_letter(ord(col_b) - ord('A') + 1)}{row - 7}"
                outflow_cell = f"{get_column_letter(ord(col_b) - ord('A') + 1)}{row - 1}"
                ws[col_b] = f"={inflow_cell}-{outflow_cell}"
                ws[col_c] = f"={inflow_cell.replace(col_b, col_c)}-{outflow_cell.replace(col_b, col_c)}"
            elif "筹资" in item:
                inflow_cell = f"{get_column_letter(ord(col_b) - ord('A') + 1)}{row - 7}"
                outflow_cell = f"{get_column_letter(ord(col_b) - ord('A') + 1)}{row - 1}"
                ws[col_b] = f"={inflow_cell}-{outflow_cell}"
                ws[col_c] = f"={inflow_cell.replace(col_b, col_c)}-{outflow_cell.replace(col_b, col_c)}"
            elif "净增加额" in item:
                ws[col_b] = f"=C16+C32+C46+C47"
                ws[col_c] = f"=D16+D32+D46+D47"
            elif "期末" in item:
                ws[col_b] = f"=C48+C49"
                ws[col_c] = f"=D48+D49"
    row += 1

# 设置数字格式
for col in ['B', 'C']:
    for row in range(5, 51):
        cell = ws[f'{col}{row}']
        if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
            continue
        cell.number_format = '#,##0.00'

# 设置颜色
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

for row in range(5, 51):
    cell = ws[f'A{row}']
    if "流入" in cell.value or cell.value in ["销售商品、提供劳务收到的现金", "收到的税费返还", "收到其他与经营活动有关的现金", "收回投资收到的现金", "取得投资收益收到的现金", "处置固定资产...收回的现金净额", "处置子公司...收到的现金净额", "收到其他与投资活动有关的现金", "吸收投资收到的现金", "取得借款收到的现金", "发行债券收到的现金", "收到其他与筹资活动有关的现金"]:
        ws[f'B{row}'].font = Font(color="006400")
        ws[f'C{row}'].font = Font(color="006400")
    elif "流出" in cell.value or cell.value in ["购买商品、接受劳务支付的现金", "支付给职工...支付的现金", "支付的各项税费", "支付其他与经营活动有关的现金", "购建固定资产...支付的现金", "投资支付的现金", "取得子公司...支付的现金净额", "支付其他与投资活动有关的现金", "偿还债务支付的现金", "分配股利...支付的现金", "支付其他与筹资活动有关的现金"]:
        ws[f'B{row}'].font = Font(color="FF0000")
        ws[f'C{row}'].font = Font(color="FF0000")
    elif "净额" in cell.value or "净增加额" in cell.value or "期末" in cell.value:
        ws[f'B{row}'].fill = gray_fill
        ws[f'C{row}'].fill = gray_fill
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'].font = Font(bold=True)

# 设置边框
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
for row in ws['A5:D50']:
    for cell in row:
        cell.border = thin_border

# 冻结窗格
ws.freeze_panes = 'B2'

# 保存文件
wb.save("现金流量表_中国会计准则_2025版.xlsx")
print("✅ 现金流量表Excel模板已生成：现金流量表_中国会计准则_2025版.xlsx")
